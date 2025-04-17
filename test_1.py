import psycopg2
from openpyxl import load_workbook

# Database connection parameters
db_config = {
    "dbname": "postgres",
    "user": "postgres",  # Fixed typo
    "password": "postgres",
    "host": "database-1.cqle1bus5rlk.us-east-1.rds.amazonaws.com",  # Change to your database host
    "port": "5432"        # Default PostgreSQL port
}

# Path to your Excel file
xlsx_file_path = "C:/Users/myrth/Documents/GitHub/big-data-infrastructure-exercises/test.xlsx"

# Initialize the connection variable
conn = None

# Connect to the PostgreSQL database
try:
    conn = psycopg2.connect(**db_config)  # Unpack db_config dictionary
    cursor = conn.cursor()
    print("Connected to the database.")

    # Load the Excel file
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active  # Get the active sheet

    # Read the header row
    headers = [cell.value for cell in sheet[1]]

    # Insert data into the aircraft_data table
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        cursor.execute(
            f"INSERT INTO aircraft_data ({', '.join(headers)}) VALUES ({', '.join(['%s'] * len(row))})",
            row
        )

    # Commit the transaction
    conn.commit()
    print("Data inserted successfully.")

except Exception as e:
    print(f"Error: {e}")

finally:
    # Close the database connection if it was successfully created
    if conn:
        cursor.close()
        conn.close()
        print("Database connection closed.")
